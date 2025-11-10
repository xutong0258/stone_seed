from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import os
import numpy as np


def write_to_excel(filename, data, headers=None, sheet_name="Sheet1"):
    """
    使用openpyxl向Excel文件写入数据

    参数:
        filename (str): 要写入的Excel文件名（.xlsx格式）
        data (list): 要写入的数据，格式为列表的列表
        headers (list, optional): 表头列表
        sheet_name (str, optional): 工作表名称
    """
    try:
        # 创建工作簿
        wb = Workbook()
        # 获取活动工作表
        ws = wb.active
        # 设置工作表名称
        ws.title = sheet_name

        # 写入表头（如果提供）
        if headers:
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                # 设置表头样式：加粗
                cell.font = Font(bold=True)
                # 居中对齐
                cell.alignment = Alignment(horizontal="center")

        # 写入数据
        start_row = 2 if headers else 1
        for row, data_row in enumerate(data, start_row):
            for col, value in enumerate(data_row, 1):
                ws.cell(row=row, column=col).value = value

        # 保存文件
        wb.save(filename)
        print(f"成功将数据写入到 {os.path.abspath(filename)}")
        return True

    except Exception as e:
        print(f"写入Excel时发生错误: {str(e)}")
        return False


if __name__ == "__main__":
    # 示例数据
    sample_data = [
        ["张三", 25, "工程师", "北京", 8000],
        ["李四", 30, "设计师", "上海", 9500],
        ["王五", 35, "产品经理", "广州", 12000],
        ["赵六", 28, "开发工程师", "深圳", 10000]
    ]

    # 表头
    headers = ["姓名", "年龄", "职业", "城市", "月薪"]

    # 写入Excel文件
    write_to_excel("员工信息.xlsx", sample_data, headers, "员工数据")
